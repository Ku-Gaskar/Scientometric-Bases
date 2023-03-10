import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import styles as st
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd

from tkinter import filedialog
import tkinter.messagebox as mb
import sys
from datetime import date
import os.path
import re

from bs4 import BeautifulSoup as soup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC   

#Experimental and numerical analysis of mechanical characteristics of fused deposition processed honeycomb fabricated from PLA or ULTEM 9085
      
HNURE_aff="Kharkiv National University of Radio Electronics"
                           
Login='olha.kusmarova@nure.ua'
Pas='jxtgznj4rf!'

mas_Doc=[4.0,40.0,25.0,35.0,6.0,11.0,28.0,20.0]
mas_Doc_Name=['№','Title Document','Autors','Autors HNURE','Kaf','Journal','Error','Error Data']

mas_Dep=[3.0,55.0,35.0,4.6,40.0]
mas_Dep_Name=[' ','Deport/Title Document','Autors HNURE','SUM','Journal']

bd = Side(style='thick', color="000000")
bl = Side(style='thin', color="000000")

def Find_exel(sheet_active,diap,search_text):
    
    abcd=openpyxl.utils.cell.range_boundaries(diap)
    res=[]
    search_text=search_text.strip()
    for  column_min in range(abcd[0],abcd[2]+1):
        for row_min_min in range(abcd[1],abcd[3]+1):
            word_cell = str(get_column_letter(column_min)) + str(row_min_min)
            data_from_cell = str(sheet_active[word_cell].value)
            if search_text in data_from_cell:
                res.append(word_cell)
            '''
            for name in data_from_cell.split(';'):
                if search_text == name.strip():
                    res.append(word_cell)
                    break
            '''          
    return res

def set_Width(ws,d):
    for i,val in enumerate(d):
        ws.column_dimensions[get_column_letter(i+1)].width = val

def set_Header(ws,i_row,Doc_Name):
     for i_col,val in enumerate(Doc_Name):
        adr=get_column_letter(i_col+1)+str(i_row)
        ws[adr].value=val
        ws[adr].alignment =Alignment(horizontal='center',vertical='center')
        ws[adr].font=Font(bold=True)
        ws[adr].border=Border(left=bd, top=bd, right=bd, bottom=bd)

def view_Error(ws,i,text,dateError):
    ws['G'+str(i)].value=text#'Error: Autor HNURE not found'
    ws['H'+str(i)].value=dateError
    for tyty in ws['A'+ str(i)+':H'+ str(i)]:
        for strk in tyty:
            strk.fill =PatternFill('solid',fgColor='FFFF00')

def search_HNURE(c,ar):
    for l in c:
        for h in HNURE_aff:
            univer=l.text.lower()
            if h in univer and ('kharkiv' in univer or 'kharkov' in univer or 'kharkіv' in univer):
                ar.append(l.text[0])
    return ar  

def comma_in_Name(Name):
    if not Name : return Name
    s=Name.split(' ')
    if s[0][-1]==',': return Name
    return s[0] +', '+''.join(s[i] for i in range(1,len(s)))

def autorized_WOS():
    driver.implicitly_wait(10)
    driver.get('https://www.webofscience.com/wos/woscc/basic-search')
    wait=WebDriverWait(driver,8)
    try:
        wait.until(EC.visibility_of_element_located((By.TAG_NAME,'app-search-home')))
    except:
        w=driver.find_element(By.CSS_SELECTOR,'login.ng-star-inserted').find_elements(By.CSS_SELECTOR,'div.mat-form-field-infix')
        w1=w[0].find_element(By.TAG_NAME,'input')
        w1.click()
        w1.send_keys(Login)
        w2=w[1].find_element(By.TAG_NAME,'input')
        w2.click()
        w2.send_keys(Pas)
        driver.find_element(By.ID,'signIn-btn').click()
    
    try:
        WebDriverWait(driver,15).until(EC.visibility_of_element_located((By.CSS_SELECTOR,'div.search-form.search-form-margin'))) 
        return True    
    except:
        return False  

def search_Header(S):
    for i in range(3,S.nrows):
        if S.cell_value(rowx=i, colx=0)=='Title':
            return i
    return -1
    
def search_collum_Header(S,row,text):
    for i in range(1,S.ncols):
        if S.cell_value(rowx=row, colx=i)==text:
            return i
    return -1

def search_Article(a_T): #ищет статью если находит то загружает ее и возвращает ИСТИНА
    w=driver.find_element(By.TAG_NAME,'app-search-row')
    sel=w.text.split('\n')
    if not (sel[0] == 'Title' or sel[0] =='Заголовок публикации'):
        w.find_element(By.CSS_SELECTOR,'button.dropdown').click()
        w1=w.find_element(By.CSS_SELECTOR,'div.options').find_elements(By.XPATH,'//div[@tabindex="-1"]')
        w1[4].click()
    driver.implicitly_wait(0.2)
    if len(driver.find_elements(By.CSS_SELECTOR,'button.clear-row-button.ng-star-inserted'))>0:
        driver.find_element(By.CSS_SELECTOR,'button.clear-row-button.ng-star-inserted').click()
    driver.implicitly_wait(10)
    w2=w.find_element(By.XPATH,'//input[@name="search-main-box"]')
    w2.click()
    w2.send_keys(a_T)
    driver.find_element(By.CSS_SELECTOR,'div.button-row').find_elements(By.TAG_NAME,'button')[1].click()
    w=driver.find_element(By.TAG_NAME,'app-general-search-friendly-display').find_element(
        By.CSS_SELECTOR,'span.brand-blue').text
    if int(w) > 0 : 
        w1=driver.find_element(By.CSS_SELECTOR,'div.data-section').find_element(By.XPATH,'//app-summary-title/h3/a').click()
        try:
            WebDriverWait(driver,10).until(EC.visibility_of_any_elements_located((By.TAG_NAME,'app-full-record-author-organization')))
            return True
        except:
            pass
    print(a_T+'\nНайдено:'+w) 
    return False

def read_Affiletions(H_af): #ищет аффилиацию ХИРЕ и возвращает массив цифр в противном []
    res=[]
    t1=so.find_all('app-full-record-author-organization')
    for a in t1:
        c=a.get_text()
        if ('Аффилиация' in c and H_af in c) or ('Affiliation' in c and H_af in c):
            try:
                res.append(re.findall('^arrow_drop_down([\d]+)',c)[0]) 
            except IndexError:
                continue
#    if not res:
#        print (res)

    return res        

def read_Autors():#возвращает список авторов с аффилиациями в противном []
    w=so.select('span[id^="author-"]')
    res={}
    for i in w:
        if not i.text: continue
        try:
            res[re.findall('\((.*)\)',i.text)[0]]=re.findall('\[([\d]+)\]',i.text)
        except IndexError:
            continue
#    print(res)
    return res

#***************************************************************************************************

Path_GreenTable=filedialog.askopenfilename(title='Откройте \"Таблица ученных ...\"',filetypes=[('Excel file','*.xlsx')])
if not Path_GreenTable: 
  mb.showwarning("Ошибка","Файл зеленой таблицы не выбран. \n Программа завершит работу")
  sys.exit()

GreenTable = openpyxl.load_workbook(Path_GreenTable)
Grt=GreenTable.active


Path_SavedrecTable=filedialog.askopenfilename(title='Откройте \"savedrecs*.xls\"',filetypes=[('savedrecs.xls file','savedrecs*.xls')])
if not Path_SavedrecTable: 
    mb.showwarning("Ошибка","Файл savedrecs*.xls не выбран. \n Программа завершит работу")
    sys.exit()

file_name=re.findall('([\s\w()\-_]*\.xls)',Path_SavedrecTable)[0]
SavedrecTable = xlrd.open_workbook(Path_SavedrecTable)
Svd=SavedrecTable.sheet_by_index(0)

row_Header=search_Header(Svd)
if row_Header == -1:
    mb.showwarning("Ошибка","Несоответствие формата savedrecs*.xls  \n Программа завершит работу")
    sys.exit()
col_Autors=search_collum_Header(Svd,row_Header,'Authors')
col_Jornal=search_collum_Header(Svd,row_Header,'Source Title')

if col_Autors == -1 or col_Jornal== -1:
    mb.showwarning("Ошибка","Несоответствие формата savedrecs*.xls  \n Программа завершит работу")
    sys.exit()

GreenTable.close()

wb = openpyxl.Workbook()
SheetAc=wb.active
SheetAc.title='Documents'
Sheet_out=wb.create_sheet('Kafedry')
i_SheetAc=1
set_Width(SheetAc,mas_Doc)
set_Header(SheetAc,i_SheetAc,mas_Doc_Name)
SheetAc.freeze_panes='A2'
i_SheetAc+=1
kaf=[]

chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument('--user-data-dir=C:\\Users\\Professional\\AppData\\Local\\Google\\Chrome\\User Data\\Profile 3')
driver = webdriver.Chrome(chrome_options=chrome_options)

i_count_Doc=0

for i in range(row_Header+1,Svd.nrows):
    if not autorized_WOS(): 
        mb.showwarning("Ошибка"," Авторизация не прошла. \n Программа завершит работу")
        driver.close()
        sys.exit()
    i_count_Doc+=1
    a_Autors=Svd.cell_value(rowx=i, colx=col_Autors)
    a_Autors=a_Autors.split('; ')
    a_Title=Svd.cell_value(rowx=i, colx=0)
    a_Jornal=Svd.cell_value(rowx=i, colx=col_Jornal)

    SheetAc['A'+str(i_SheetAc)].value=i_count_Doc
    SheetAc['B'+str(i_SheetAc)].value=a_Title
    
    if not search_Article(a_Title):
        mb.showwarning("Ошибка","Статья не найдена:\n%s\n Программа завершит работу"%a_Title)
        print("Ошибка: Статья не найдена->"+a_Title)
        driver.close()
        sys.exit()

    data=driver.page_source
    so=soup(data,'lxml')
    HNURE_char=read_Affiletions(HNURE_aff)

    if not HNURE_char:
        view_Error (SheetAc,i_SheetAc,'Error: Not affiletions HNURE',' ') 
    autors=read_Autors()     
        
    for a_One,affil in autors.items():
        at_Struct=[]            
        at_Struct.append(a_Title)
        SheetAc['C'+str(i_SheetAc)].value=a_One
        for l in affil:
            if l in HNURE_char:
                addr=Find_exel(Grt,'R3:R'+str(Grt.max_row),a_One)
                if len(addr)==1:
                    sdf=openpyxl.utils.cell.coordinate_from_string(addr[0])
                    name=Grt['B'+str(sdf[1])].value
                    at_Struct.append(name)
                    SheetAc['D'+str(i_SheetAc)].value=name
                    n_kaf=Grt['L'+str(sdf[1])].value
                    if n_kaf :
                        n_kaf=n_kaf.upper() 
                    else:
                        n_kaf='???'
                    SheetAc['E'+str(i_SheetAc)].value=n_kaf
                    SheetAc['F'+str(i_SheetAc)].value=a_Jornal
                    at_Struct.append(n_kaf)
                    at_Struct.append(a_Jornal)
                    kaf.append(at_Struct)
                elif len(addr)>1:
                    SheetAc['G'+str(i_SheetAc)].value='Error: Found multiple authors'
                    for s in addr:
                        sdf=openpyxl.utils.cell.coordinate_from_string(s)
                        n_kaf=Grt['L'+str(sdf[1])].value
                        if n_kaf :
                            n_kaf=n_kaf.upper() 
                        else:
                            n_kaf='???'
                        SheetAc['H'+str(i_SheetAc)].value=a_One+'->'+ Grt['B'+str(sdf[1])].value+'->'+n_kaf 
                        i_SheetAc+=1
                    i_SheetAc-=1
                    for  tyty in SheetAc['A'+str(i_SheetAc-len(addr)):'H'+ str(i_SheetAc)]:
                        for strk in tyty :#cell_range:  E-G
                            strk.fill =st.PatternFill('solid',fgColor='FFFF00') 
                    print('Найдено несколько  записей ->'+a_One)        
                else:
                    view_Error (SheetAc,i_SheetAc,'Error: Autor HNURE not found',a_One)                
                    print('Aвтор %s не найден' %a_One)
        i_SheetAc+=1
      
for i,t in enumerate(SheetAc['B']):
    #for strk in tyty:
        if i>0 and t.value: 
            for x in range(1,9):
                SheetAc[get_column_letter(x)+str(i+1)].border=Border(top=bl)
for x in range(1,9):
    SheetAc[get_column_letter(x)+str(SheetAc.max_row)].border=Border(bottom=bl)
  
y_Sheet_out=1
set_Width(Sheet_out,mas_Dep)
set_Header(Sheet_out,y_Sheet_out,mas_Dep_Name)
Sheet_out.freeze_panes='A2'
list_kaf=sorted(set(kaf_i[2] for kaf_i in kaf))
for i_k in list_kaf:
    y_Sheet_out+=1
    Sheet_out['B'+str(y_Sheet_out)].value='Кафедра \"'+i_k+'\"'
    Sheet_out['B'+str(y_Sheet_out)].font=Font(bold=True)

    l_ind=y_Sheet_out
    count=0
    y_Sheet_out+=1
    list_at=[]
    for j in kaf:
        if i_k==j[2]:
            if j[0] in list_at: continue
            Sheet_out['B'+str(y_Sheet_out)].value=j[0]
            Sheet_out['C'+str(y_Sheet_out)].value=j[1]
            Sheet_out['E'+str(y_Sheet_out)].value=j[3]
            list_at.append(j[0])
            count+=1
            y_Sheet_out+=1
    Sheet_out['D'+str(l_ind)].value=count
    Sheet_out['D'+str(l_ind)].font=Font(bold=True)
current_date = str(date.today())
Namefile='Result_'+current_date+'.xlsx'
i=1
while os.path.isfile(Namefile):
    Namefile='Result'+current_date+'('+str(i)+')'+'.xlsx'
    i+=1
driver.close()
wb.save(Namefile)
wb.close()  
sys.exit()