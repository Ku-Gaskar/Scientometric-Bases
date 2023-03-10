import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import tkinter.messagebox as mb
import sys
from datetime import date
import os.path
import re
import PySimpleGUI as sg
import xlrd

def Find_exel(sheet_active,diap,search_text):
    abcd=openpyxl.utils.cell.range_boundaries(diap)
    for  column_min in range(abcd[0],abcd[2]+1):
        for row_min_min in range(abcd[1],abcd[3]+1):
            word_cell = str(get_column_letter(column_min)) + str(row_min_min)
            data_from_cell = str(sheet_active[word_cell].value)
            if len(re.findall(search_text, data_from_cell)) > 0:
                return word_cell
    return False

#*****************************************************************************    
Path_GreenTable=filedialog.askopenfilename(title='Откройте \"Таблица ученных ...\"',filetypes=[('Excel file','*.xlsx')])
if not Path_GreenTable: 
  mb.showwarning("Ошибка","Файл зеленой таблицы не выбран. \n Программа завершит работу")
  sys.exit()

GreenTable = openpyxl.load_workbook(Path_GreenTable)
Grt=GreenTable.active


Path_SavedrecTable=filedialog.askopenfilename(title='Откройте \"savedrecs*.xls\"',filetypes=[('savedrecs.xls file','savedrecs*.xls')])
if not Path_SavedrecTable: 
  mb.showwarning("Ошибка","Файл savedrecs*.xls не выбран. \n Программа завершит работу")
  sys.exit()

file_name=re.findall('([\s\w()\-_]*\.xls)',Path_SavedrecTable)[0]


SavedrecTable = xlrd.open_workbook(Path_SavedrecTable)
Svd=SavedrecTable.sheet_by_index(0)

#SavedrecTable.close()
GreenTable.close()

wb = openpyxl.Workbook()
SheetAc=wb.active
SheetAc.title='Documents'
Sheet_out=wb.create_sheet('Kafedry')
i_SheetAc=2

kaf=[]
#tt=Find_exel(Svd,'A1:A20','Title')
#if not tt:
#    mb.showwarning("Ошибка","Формат файла savedrecs*.xls не соответствует.\n Программа завершит работу")
#    sys.exit() 
#t_row=openpyxl.utils.cell.coordinate_from_string(tt)[1]
for i in range(11,Svd.nrows):
    sg.one_line_progress_meter('Обработка: '+ file_name,i,max_value=Svd.nrows, no_button=False)
    a_Autors=Svd.cell_value(rowx=i, colx=1)
    a_Autors=a_Autors.split('; ')
    a_Title=Svd.cell_value(rowx=i, colx=0)
    a_Jornal=Svd.cell_value(rowx=i, colx=5)
    SheetAc['B'+str(i_SheetAc)].value=a_Title
    for a_One in a_Autors:
            at_Struct=[]            
            at_Struct.append(a_Title)
            SheetAc['C'+str(i_SheetAc)].value=a_One
            addr=Find_exel(Grt,'R3:R'+str(Grt.max_row),a_One)
            if addr:
                sdf=openpyxl.utils.cell.coordinate_from_string(addr)
                name=Grt['B'+str(sdf[1])].value
                at_Struct.append(name)
                SheetAc['D'+str(i_SheetAc)].value=name
                n_kaf=Grt['L'+str(sdf[1])].value
                if n_kaf :
                    n_kaf=n_kaf.upper() 
                else:
                    n_kaf='____'
                SheetAc['E'+str(i_SheetAc)].value=n_kaf
                SheetAc['F'+str(i_SheetAc)].value=a_Jornal
                at_Struct.append(n_kaf)
                at_Struct.append(a_Jornal)
                kaf.append(at_Struct)
            i_SheetAc+=1
y_Sheet_out=1
list_kaf=sorted(set(kaf_i[2] for kaf_i in kaf))
for i_k in list_kaf:
    y_Sheet_out+=1
    Sheet_out['B'+str(y_Sheet_out)].value='Кафедра \"'+i_k+'\"'
    l_ind=y_Sheet_out
    count=0
    y_Sheet_out+=1
    list_at=[]
    for j in kaf:
        if i_k==j[2]:
            if j[0] in list_at: continue
            Sheet_out['B'+str(y_Sheet_out)].value=j[0]
            Sheet_out['C'+str(y_Sheet_out)].value=j[1]
            Sheet_out['E'+str(y_Sheet_out)].value=j[3]
            list_at.append(j[0])
            count+=1
            y_Sheet_out+=1
    Sheet_out['D'+str(l_ind)].value=count
sg.one_line_progress_meter_cancel()
current_date = str(date.today())
Namefile='Result_'+current_date+'.xlsx'
i=1
while os.path.isfile(Namefile):
    Namefile='Result'+current_date+'('+str(i)+')'+'.xlsx'
    i+=1
SavedrecTable.release_resources()
del SavedrecTable
wb.save(Namefile)
wb.close()  