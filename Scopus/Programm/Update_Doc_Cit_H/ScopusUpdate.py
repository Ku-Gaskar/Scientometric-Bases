import sys
from time import sleep
from tkinter import filedialog
import tkinter.messagebox as mb
import openpyxl
from easygui import integerbox
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import styles as st
from bs4 import BeautifulSoup
from employee_class import emp
from datetime import date

RedColor='FF0000'

Path_GreenTable=filedialog.askopenfilename(title='Откройте \"Таблица ученных ...\"',filetypes=[('Excel file','*.xlsx')])
if not Path_GreenTable: 
  mb.showwarning("Ошибка","Файл зеленой таблицы не выбран. \n Программа завершит работу")
  sys.exit()

GreenTable = openpyxl.load_workbook(Path_GreenTable)
SheetAc=GreenTable.active
Start_Number=integerbox("Введите начальный номер строки","Ввод числа от 3 до "+ str(SheetAc.max_row),None,3,SheetAc.max_row)
CountRepead=integerbox("Введите количество повторов","Ввод числа от 1 до "+ str(SheetAc.max_row-3),None,1,SheetAc.max_row-3)

fillConf='00FFFF'

for  tyty in SheetAc['E'+str(Start_Number):'G'+ str(Start_Number + CountRepead - 1)]:
  for strk in tyty :#cell_range:  E-G
    if strk.row > 2 :
         strk.fill =st.PatternFill('solid',fgColor='FFFFFF') 

driver = webdriver.Chrome() 
from selenium.webdriver.chrome.options import Options
chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
#chrome_options.add_argument('--user-data-dir=C:\\Users\\Professional\\AppData\\Local\\Google\\Chrome\\User Data\\Profile 1')
driver = webdriver.Chrome(chrome_options=chrome_options)
Client=emp

for i2 in range(Start_Number,Start_Number + CountRepead):
  url=SheetAc['D'+str(i2)].value
  if url : 
    driver.get(url)
    driver.implicitly_wait(10)  # Установить 10 секунд времени ожидания
    #sleep(1)
    data_page = driver.page_source
    soup = BeautifulSoup(data_page, 'lxml')
    TagSearch1 = soup.find('div', class_= 'col-lg-6 col-24')
    for I_21 in  range(3):
      if not TagSearch1 or TagSearch1.getText()=='' : 
        sleep(2)
        driver.implicitly_wait(10)  # Установить 10 секунд времени ожидания
        data_page = driver.page_source
        soup = BeautifulSoup(data_page, 'lxml')
        TagSearch1 = soup.find('div', class_= 'col-lg-6 col-24')
      else: break
    if not  TagSearch1 :
        print('Данных нет')
        continue 

    TagSearch = TagSearch1.find_all('div')
    if not  TagSearch :
        mb.showinfo('Ошибка данных','Изменилась cтраница Scopus!!!')
        driver.quit()
        sys.exit()
    Client.CountDoc= int(TagSearch[0].next.getText())
    Client.Citir = int(TagSearch[1].next.getText())
    Client.H_ind = int(TagSearch[2].next.getText())
    
    if int(SheetAc['E'+str(i2)].value) <= Client.CountDoc : SheetAc['E'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf) 
    else : SheetAc['E'+str(i2)].fill=st.PatternFill('solid',fgColor=RedColor) 
    if int(SheetAc['F'+str(i2)].value) <= Client.Citir : SheetAc['F'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf) 
    else : SheetAc['F'+str(i2)].fill=st.PatternFill('solid',fgColor=RedColor)
    if int(SheetAc['G'+str(i2)].value) <= Client.H_ind: SheetAc['G'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf) 
    else : SheetAc['G'+str(i2)].fill=st.PatternFill('solid',fgColor=RedColor) 
    SheetAc['E'+str(i2)].value=Client.CountDoc
    SheetAc['E'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf)
    SheetAc['F'+str(i2)].value=Client.Citir
    SheetAc['G'+str(i2)].value=Client.H_ind
current_date = str(date.today())
Namefile='Таблица ученных '+current_date+'.xlsx'
GreenTable.save(Namefile)
GreenTable.close()  
driver.quit()
mb.showinfo('Программа завершена','Результат в файле:\n'+Namefile)
sys.exit()