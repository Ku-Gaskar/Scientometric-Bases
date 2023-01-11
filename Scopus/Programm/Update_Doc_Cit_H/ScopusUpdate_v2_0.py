import sys
import os
import re

from tkinter import filedialog
import tkinter.messagebox as mb
from easygui import integerbox

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

import openpyxl
from openpyxl import styles as st
from employee_class import Emp

from datetime import date

#**************************** functions *****************************************
def Close_Aplications():
    current_date = str(date.today())
    Namefile=re.sub('[\(\)\-_0-9]*(.xlsx)',current_date+r'\1',Path_GreenTable,count=1)
    i=1
    Name_=Namefile[:-5]
    while os.path.isfile(Namefile):
      Namefile=Name_+'('+str(i)+')'+'.xlsx'
      i+=1

    GreenTable.save(Namefile)
    GreenTable.close()  
    driver.quit()
    mb.showinfo('Программа завершена','Результат в файле:\n'+Namefile)
    sys.exit()

#******************************* main ******************************************
if __name__=='__main__':
    
  RedColor='FF0000'
  fillConf='00FFFF'

  Path_GreenTable=filedialog.askopenfilename(title='Откройте \"Таблица ученных ...\"',filetypes=[('Excel file','*.xlsx')])
  if not Path_GreenTable: 
    mb.showwarning("Ошибка","Файл зеленой таблицы не выбран. \n Программа завершит работу")
    sys.exit()

  GreenTable = openpyxl.load_workbook(Path_GreenTable)
  SheetAc=GreenTable.active
  Start_Number=integerbox("Введите начальный номер строки","Ввод числа от 3 до "+ str(SheetAc.max_row),3,3,SheetAc.max_row)
  CountRepead=integerbox("Введите количество повторов","Ввод числа от 1 до "+ str(SheetAc.max_row-3),SheetAc.max_row-3,1,SheetAc.max_row-3)

  for  tyty in SheetAc['E'+str(Start_Number):'G'+ str(Start_Number + CountRepead - 1)]:
    for strk in tyty :#cell_range:  E-G
      if strk.row > 2 :
          strk.fill =st.PatternFill('solid',fgColor='FFFFFF') 

  chrome_options = Options()
  chrome_options.add_argument("--disable-extensions")
  #chrome_options.add_argument('--user-data-dir=C:\\Users\\Professional\\AppData\\Local\\Google\\Chrome\\User Data\\Profile 1')
  driver = webdriver.Chrome(chrome_options=chrome_options)
  client=Emp

  for i2 in range(Start_Number,Start_Number + CountRepead):
    url=SheetAc['D'+str(i2)].value
    if url : 
      driver.implicitly_wait(20)  # Установить 20 секунд времени ожидания
      driver.get(url)
      try:
        WebDriverWait(driver,30).until(EC.visibility_of_element_located((By.ID,'highcharts-information-region-1'))) 
      except:
        mb.showwarning("Ошибка","Страница не загружена. \n Программа завершит работу")
        Close_Aplications()        

      try:
        s1=driver.find_elements(By.CSS_SELECTOR,'span.typography_ceae25.font-size-xl_ceae25.sans_ceae25')
        s2=driver.find_element(By.CSS_SELECTOR,'#scopus-author-profile-page-control-microui__documents-tab > span').text
        s2=re.findall(r'[ 0-9]*',s2)[0]
        client.Citir=int(s1[0].text.replace(' ',''))
        client.CountDoc=int(s2.replace(' ',''))
        client.H_ind=int(s1[2].text.replace(' ',''))
      except:  
        mb.showwarning("Ошибка","Ошибка на странице. \n Программа завершит работу")
        Close_Aplications()       

      #print('doc = {}, {}, {}'.format(client.CountDoc,client.Citir,client.H_ind)) 
      if SheetAc['E'+str(i2)].value == None: SheetAc['E'+str(i2)].value = 0
      if int(SheetAc['E'+str(i2)].value) <= client.CountDoc : SheetAc['E'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf) 
      else : SheetAc['E'+str(i2)].fill=st.PatternFill('solid',fgColor=RedColor) 
      
      if SheetAc['F'+str(i2)].value == None: SheetAc['F'+str(i2)].value = 0
      if int(SheetAc['F'+str(i2)].value) <= client.Citir : SheetAc['F'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf) 
      else : SheetAc['F'+str(i2)].fill=st.PatternFill('solid',fgColor=RedColor)
      
      if SheetAc['G'+str(i2)].value == None: SheetAc['G'+str(i2)].value = 0
      if int(SheetAc['G'+str(i2)].value) <= client.H_ind: SheetAc['G'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf) 
      else : SheetAc['G'+str(i2)].fill=st.PatternFill('solid',fgColor=RedColor) 
      SheetAc['E'+str(i2)].value=client.CountDoc
      SheetAc['E'+str(i2)].fill=st.PatternFill('solid',fgColor=fillConf)
      SheetAc['F'+str(i2)].value=client.Citir
      SheetAc['G'+str(i2)].value=client.H_ind
  
  Close_Aplications()