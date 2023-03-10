import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import styles as st
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from tkinter import filedialog
import tkinter.messagebox as mb
import sys
from datetime import date
import os.path
import re
#import PySimpleGUI as sg
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
#from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC   

from time import sleep

        
HNURE=['national university of radio electronics',
       'national university of radioelectronics',
       'national university radioelectronics',
       'national, university of radio electronics']

Login='olha.kusmarova@nure.ua'
Pas='Jkm4br1!'

mas_Doc=[4.0,40.0,25.0,35.0,6.0,11.0,28.0,20.0]
mas_Doc_Name=['№','Title Document','Autors','Autors HNURE','Kaf','Journal','Error','Error Data']

mas_Dep=[3.0,55.0,35.0,4.6,40.0]
mas_Dep_Name=[' ','Deport/Title Document','Autors HNURE','SUM','Journal']

bd = Side(style='thick', color="000000")
bl = Side(style='thin', color="000000")

def Find_exel(sheet_active,diap,search_text):
    
    abcd=openpyxl.utils.cell.range_boundaries(diap)
    res=[]
    search_text=search_text.strip()
    for  column_min in range(abcd[0],abcd[2]+1):
        for row_min_min in range(abcd[1],abcd[3]+1):
            word_cell = str(get_column_letter(column_min)) + str(row_min_min)
            data_from_cell = str(sheet_active[word_cell].value)
            if search_text in data_from_cell:
                res.append(word_cell)
            '''
            for name in data_from_cell.split(';'):
                if search_text == name.strip():
                    res.append(word_cell)
                    break
            '''          
    return res

def set_Width(ws,d):
    for i,val in enumerate(d):
        ws.column_dimensions[get_column_letter(i+1)].width = val

def set_Header(ws,i_row,Doc_Name):
     for i_col,val in enumerate(Doc_Name):
        adr=get_column_letter(i_col+1)+str(i_row)
        ws[adr].value=val
        ws[adr].alignment =Alignment(horizontal='center',vertical='center')
        ws[adr].font=Font(bold=True)
        ws[adr].border=Border(left=bd, top=bd, right=bd, bottom=bd)

def view_Error(ws,i,text,dateError):
    ws['G'+str(i)].value=text#'Error: Autor HNURE not found'
    ws['H'+str(i)].value=dateError
    for tyty in ws['A'+ str(i)+':H'+ str(i)]:
        for strk in tyty:
            strk.fill =PatternFill('solid',fgColor='FFFF00')

def search_HNURE(c,ar):
    for l in c:
        for h in HNURE:
            univer=l.text.lower()
            if h in univer and ('kharkiv' in univer or 'kharkov' in univer or 'kharkіv' in univer):
                ar.append(l.text[0])
    return ar  

def comma_in_Name(Name):
    if not Name : return Name
    s=Name.split(' ')
    if s[0][-1]==',': return Name
    return s[0] +', '+''.join(s[i] for i in range(1,len(s)))


def autorizedScopus():
    try:
        driver.get('https://www.scopus.com')
        cc=driver.find_element(By.CSS_SELECTOR,'div.GlobalHeader-module__1G1hI').find_elements(By.TAG_NAME,'title')
        if cc[1].get_attribute('id') !='gh-wm-scopus':
            driver.get('https://www.scopus.com/signin.uri?origin=&zone=TopNavBar')
            d=driver.find_element(By.CSS_SELECTOR,'div.form-row').find_element(By.ID,'bdd-email')
            d.click()
            d.send_keys(Login)
            driver.find_element(By.ID,'bdd-elsPrimaryBtn').click()
            sleep(1)
            driver.find_element(By.ID,'bdd-password').send_keys(Pas)
            driver.find_element(By.ID,'bdd-elsPrimaryBtn').click()
        return True
    except:
        return False
#***************************************************************************************************

Path_GreenTable=filedialog.askopenfilename(title='Откройте \"Таблица ученных ...\"',filetypes=[('Excel file','*.xlsx')])
if not Path_GreenTable: 
  mb.showwarning("Ошибка","Файл зеленой таблицы не выбран. \n Программа завершит работу")
  sys.exit()

GreenTable = openpyxl.load_workbook(Path_GreenTable)
Grt=GreenTable.active

Path_file_bib=filedialog.askopenfilename(title='Откройте \"scopus...bib\"',filetypes=[('bib file','*.bib')])
if not Path_file_bib: 
  mb.showwarning("Ошибка","Файл *.bib не выбран. \n Программа завершит работу")
  sys.exit()

Bib_file = open(Path_file_bib,'r',encoding='UTF-8')
buff=Bib_file.read()
Bib_file.close()
GreenTable.close()

wb = openpyxl.Workbook()
SheetAc=wb.active
SheetAc.title='Documents'
Sheet_out=wb.create_sheet('Kafedry')
i_SheetAc=1
set_Width(SheetAc,mas_Doc)
set_Header(SheetAc,i_SheetAc,mas_Doc_Name)
SheetAc.freeze_panes='A2'
i_SheetAc+=1
kaf=[]

At=buff.split('@')

chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument('--user-data-dir=C:\\Users\\Professional\\AppData\\Local\\Google\\Chrome\\User Data\\Profile 3')
driver = webdriver.Chrome(chrome_options=chrome_options)

if not autorizedScopus(): 
    mb.showwarning("Ошибка"," Авторизация не прошла. \n Программа завершит работу")
    sys.exit()

for i,atcl in enumerate(At):
    if i>0:
        a_Autors=re.findall("author=\{.*\}",atcl)[0][8:-1]
        a_Autors=a_Autors.split(' and ')
        a_Title=re.findall("title=\{.*\}",atcl)[0][7:-1]
        a_Jornal=re.findall("journal=\{.*\}",atcl)[0][9:-1]
        url_Aticl=re.findall("url=\{(.*)\}",atcl)[0]
        SheetAc['A'+str(i_SheetAc)].value=i
        SheetAc['B'+str(i_SheetAc)].value=a_Title
        
        driver.implicitly_wait(10)
        driver.get(url_Aticl)
          # Установить 10 секунд времени ожидания
        #sleep(1)
        driver.implicitly_wait(0.5)  # Установить 0.1 секунд времени ожидания
        if len(driver.find_elements(By.ID,'show-additional-authors')):
            driver.implicitly_wait(10)
            driver.find_element(By.ID,'show-additional-authors').click()

        cc=driver.find_element(By.CSS_SELECTOR,'div.col-24.col-lg-18.col-xl-16').find_elements(By.TAG_NAME,'li')
        autors=[]
        for l in cc:
            Name_av=l.find_element(By.TAG_NAME,'els-button').text
            driver.implicitly_wait(0.1)
            Name_av+='#'+''.join(q.text for q in l.find_elements(By.TAG_NAME,'sup')) #!!!!!!!!!
            autors.append(Name_av)
            driver.implicitly_wait(10)  
        c2=driver.find_element(By.ID,'affiliation-section')
        cc=c2.find_elements(By.TAG_NAME,'li')
        VUZ_ONE=False
        if len(cc)==1: 
            VUZ_ONE=True
        HNURE_char=[]
        HNURE_char=search_HNURE(cc,HNURE_char)
        
        driver.implicitly_wait(0.1)  # Установить 0.1 секунд времени ожидания
        if len(c2.find_elements(By.ID,'show-additional-affiliations')):
            driver.implicitly_wait(10)
            c2.find_element(By.ID,'show-additional-affiliations').click()
            cc=c2.find_element(By.CSS_SELECTOR,'section.collapsible-panel__content.row').find_elements(By.TAG_NAME,'li')
            HNURE_char=search_HNURE(cc,HNURE_char)
        driver.implicitly_wait(10)  # Установить 10 секунд времени ожидания
        
        if not HNURE_char:
            view_Error(SheetAc,i_SheetAc,'Error: University not found',' ')
            print('HNURE не найден в статье->'+a_Title)
            HNURE_char.append('xx')

        for a_One in autors:
            at_Struct=[]            
            at_Struct.append(a_Title)
            mas=a_One.split('#')
            mas[0]=comma_in_Name(mas[0])
            SheetAc['C'+str(i_SheetAc)].value=mas[0]
            addr=''
            for k in HNURE_char:
                if (k in mas[1]) or (VUZ_ONE and not mas[1]):
                    addr=Find_exel(Grt,'R3:R'+str(Grt.max_row),mas[0])
                    if len(addr)==1:
                        sdf=openpyxl.utils.cell.coordinate_from_string(addr[0])
                        name=Grt['B'+str(sdf[1])].value
                        at_Struct.append(name)
                        SheetAc['D'+str(i_SheetAc)].value=name
                        n_kaf=Grt['L'+str(sdf[1])].value
                        if n_kaf :
                            n_kaf=n_kaf.upper() 
                        else:
                            n_kaf='???'
                        SheetAc['E'+str(i_SheetAc)].value=n_kaf
                        SheetAc['F'+str(i_SheetAc)].value=a_Jornal
                        at_Struct.append(n_kaf)
                        at_Struct.append(a_Jornal)
                        kaf.append(at_Struct)
                    elif len(addr)>1:
                        SheetAc['G'+str(i_SheetAc)].value='Error: Found multiple authors'
                        for s in addr:
                            sdf=openpyxl.utils.cell.coordinate_from_string(s)
                            n_kaf=Grt['L'+str(sdf[1])].value
                            if n_kaf :
                                n_kaf=n_kaf.upper() 
                            else:
                                n_kaf='???'
                            SheetAc['H'+str(i_SheetAc)].value=mas[0]+'->'+ Grt['B'+str(sdf[1])].value+'->'+n_kaf 
                            i_SheetAc+=1
                        i_SheetAc-=1
                        for  tyty in SheetAc['A'+str(i_SheetAc-len(addr)):'H'+ str(i_SheetAc)]:
                            for strk in tyty :#cell_range:  E-G
                                strk.fill =st.PatternFill('solid',fgColor='FFFF00') 
                        print('Найдено несколько  записей ->'+mas[0])        
                    else:
                        view_Error (SheetAc,i_SheetAc,'Error: Autor HNURE not found',mas[0])                
                        print('Aвтор %s не найден' %mas[0])
            i_SheetAc+=1
'''
for tyty in SheetAc['C2':'H'+str(SheetAc.max_row)]:
    for strk in tyty:
        strk.border=Border(left=bl, top=bl, right=bl, bottom=bl)          
'''
for i,t in enumerate(SheetAc['B']):
    #for strk in tyty:
        if i>0 and t.value: 
            for x in range(1,9):
                SheetAc[get_column_letter(x)+str(i+1)].border=Border(top=bl)
for x in range(1,9):
    SheetAc[get_column_letter(x)+str(SheetAc.max_row)].border=Border(bottom=bl)
  
y_Sheet_out=1
set_Width(Sheet_out,mas_Dep)
set_Header(Sheet_out,y_Sheet_out,mas_Dep_Name)
Sheet_out.freeze_panes='A2'
list_kaf=sorted(set(kaf_i[2] for kaf_i in kaf))
for i_k in list_kaf:
    y_Sheet_out+=1
    Sheet_out['B'+str(y_Sheet_out)].value='Кафедра \"'+i_k+'\"'
    Sheet_out['B'+str(y_Sheet_out)].font=Font(bold=True)

    l_ind=y_Sheet_out
    count=0
    y_Sheet_out+=1
    list_at=[]
    for j in kaf:
        if i_k==j[2]:
            if j[0] in list_at: continue
            Sheet_out['B'+str(y_Sheet_out)].value=j[0]
            Sheet_out['C'+str(y_Sheet_out)].value=j[1]
            Sheet_out['E'+str(y_Sheet_out)].value=j[3]
            list_at.append(j[0])
            count+=1
            y_Sheet_out+=1
    Sheet_out['D'+str(l_ind)].value=count
    Sheet_out['D'+str(l_ind)].font=Font(bold=True)
current_date = str(date.today())
Namefile='Result_'+current_date+'.xlsx'
i=1
while os.path.isfile(Namefile):
    Namefile='Result'+current_date+'('+str(i)+')'+'.xlsx'
    i+=1
driver.close()
wb.save(Namefile)
wb.close()  
sys.exit()