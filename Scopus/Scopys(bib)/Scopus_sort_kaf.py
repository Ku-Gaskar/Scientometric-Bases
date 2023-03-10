import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import filedialog
import tkinter.messagebox as mb
import sys
from datetime import date
import os.path
import re
import PySimpleGUI as sg



def Find_exel(sheet_active,diap,search_text):
    
    abcd=openpyxl.utils.cell.range_boundaries(diap)
 
    for  column_min in range(abcd[0],abcd[2]+1):
        for row_min_min in range(abcd[1],abcd[3]+1):
            word_cell = str(get_column_letter(column_min)) + str(row_min_min)
            data_from_cell = str(sheet_active[word_cell].value)
            if len(re.findall(search_text, data_from_cell)) > 0:
                return word_cell
            
    return False
    



Path_GreenTable=filedialog.askopenfilename(title='Откройте \"Таблица ученных ...\"',filetypes=[('Excel file','*.xlsx')])
if not Path_GreenTable: 
  mb.showwarning("Ошибка","Файл зеленой таблицы не выбран. \n Программа завершит работу")
  sys.exit()

GreenTable = openpyxl.load_workbook(Path_GreenTable)
Grt=GreenTable.active


Path_file_bib=filedialog.askopenfilename(title='Откройте \"scopus...bib\"',filetypes=[('bib file','*.bib')])
if not Path_file_bib: 
  mb.showwarning("Ошибка","Файл *.bib не выбран. \n Программа завершит работу")
  sys.exit()

Bib_file = open(Path_file_bib,'r',encoding='UTF-8')
buff=Bib_file.read()
Bib_file.close()
GreenTable.close()

wb = openpyxl.Workbook()
SheetAc=wb.active
SheetAc.title='Documents'
Sheet_out=wb.create_sheet('Kafedry')
i_SheetAc=2

kaf=[]

At=buff.split('@')
for i,atcl in enumerate(At):
    sg.one_line_progress_meter('This is my progress meter!', i+1, len(At), '-key-')
    if i>0:
        a_Autors=re.findall("author=\{.*\}",atcl)[0][8:-1]
        a_Autors=a_Autors.split(' and ')
        a_Title=re.findall("title=\{.*\}",atcl)[0][7:-1]
        a_Jornal=re.findall("journal=\{.*\}",atcl)[0][9:-1]
        SheetAc['B'+str(i_SheetAc)].value=a_Title
        
        for a_One in a_Autors:
            at_Struct=[]            
            at_Struct.append(a_Title)
            SheetAc['C'+str(i_SheetAc)].value=a_One
            addr=Find_exel(Grt,'R3:R'+str(Grt.max_row),a_One)
            if addr:
                sdf=openpyxl.utils.cell.coordinate_from_string(addr)
                name=Grt['B'+str(sdf[1])].value
                at_Struct.append(name)
                SheetAc['D'+str(i_SheetAc)].value=name
                n_kaf=Grt['L'+str(sdf[1])].value
                if n_kaf :
                    n_kaf=n_kaf.upper() 
                else:
                    n_kaf='____'
                SheetAc['E'+str(i_SheetAc)].value=n_kaf
                SheetAc['F'+str(i_SheetAc)].value=a_Jornal
                at_Struct.append(n_kaf)
                at_Struct.append(a_Jornal)
                kaf.append(at_Struct)
            i_SheetAc+=1
y_Sheet_out=1
list_kaf=sorted(set(kaf_i[2] for kaf_i in kaf))
for i_k in list_kaf:
    y_Sheet_out+=1
    Sheet_out['B'+str(y_Sheet_out)].value='Кафедра \"'+i_k+'\"'
    l_ind=y_Sheet_out
    count=0
    y_Sheet_out+=1
    list_at=[]
    for j in kaf:
        if i_k==j[2]:
            if j[0] in list_at: continue
            Sheet_out['B'+str(y_Sheet_out)].value=j[0]
            Sheet_out['C'+str(y_Sheet_out)].value=j[1]
            Sheet_out['E'+str(y_Sheet_out)].value=j[3]
            list_at.append(j[0])
            count+=1
            y_Sheet_out+=1
    Sheet_out['D'+str(l_ind)].value=count

current_date = str(date.today())
Namefile='Result_'+current_date+'.xlsx'
i=1
while os.path.isfile(Namefile):
    Namefile='Result'+current_date+'('+str(i)+')'+'.xlsx'
    i+=1
wb.save(Namefile)
wb.close()  